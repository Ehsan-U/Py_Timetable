import json
import argparse
import openpyxl
from hashlib import md5
from rich.pretty import pprint
import tkinter

class TimeTable():
    def __init__(self):
        self.hashes = {}
        self.conflict = {'Conflicts':[]}
        self.weekly = {}

    def start(self,filename):
        self.wb = openpyxl.load_workbook(filename)
        sheetnames = self.wb.sheetnames
        for sheet in sheetnames:
            wb = self.wb[sheet]
            self.get_data(wb)
            self.evaluate(wb,sheet)
            self.hashes.clear()
        # pprint(self.conflict,expand_all=True)
        with open("conflicts.json", 'w') as f:
            json.dump(self.conflict, f)
        return self.conflict

    def argss(self):
        parser = argparse.ArgumentParser()
        parser.add_argument('-f','--file',dest='filename',help='Give the excel file path',required=True)
        values = parser.parse_args()
        args_Dict = vars(values)
        return args_Dict

    def get_data(self,wb):
        for row in wb.iter_rows():
            for cell in row:
                if cell.value and cell.row > 5:
                    if cell.column == 1:
                        day = cell.value
                        self.weekly[day] = []
                    if cell.column == 5:
                        # print(cell.value,row[5].value,row[6].value)
                        teacher = cell.value
                        timing = row[5].value
                        room = row[6].value
                        self.weekly[day].append([str(day).lower(), str(teacher).lower(), str(timing).lower(), str(room).lower()])

    def evaluate(self,wb,sheet):
        # two teaches at one place
        for day, val in self.weekly.items():
            if val:
                for v in val:
                    dtr = v[0]
                    dtr += v[2]
                    dtr += v[3]
                    unique_id = md5(dtr.encode()).hexdigest()
                    if unique_id not in self.hashes:
                        self.hashes[unique_id] = v[1]
                    else:
                        self.conflict['Conflicts'].append([self.hashes.get(unique_id),v[1],v[2],v[0],sheet])
        # one teacher at two places
        for day, val in self.weekly.items():
            if val:
                for v in val:
                    dtr = v[0]
                    dtr += v[1]
                    dtr += v[2]
                    unique_id = md5(dtr.encode()).hexdigest()
                    if unique_id not in self.hashes:
                        self.hashes[unique_id] = f"{v[2]} ({v[3]})"
                    else:
                        self.conflict['Conflicts'].append([self.hashes.get(unique_id),v[1],f"{v[2]} ({v[3]})",v[0],sheet])
    def gui(self):
        root = tkinter.Tk()
        root.title('Py-resolver')
        canvass = tkinter.Canvas(root,width=400,height=300)
        canvass.pack()
        # input label
        label = tkinter.Label(master=root,text="Excel Filepath")
        canvass.create_window(200,110,window=label)
        # input box
        entry = tkinter.Entry(root,width=30)
        canvass.create_window(200,140,window=entry)
        # any func
        def lets_try():
            name = entry.get()
            if self.start(name):
                l = tkinter.Label(text='File Saved >> conflicts.json',master=root)
                canvass.create_window(200,220,window=l)
        # set button
        button = tkinter.Button(text="Submit",command=lets_try)
        canvass.create_window(200,180,window=button)
        
        root.mainloop()

# t = TimeTable()
# # args_Dict = t.argss()
# # filename = args_Dict.get("filename")
# # t.start(filename)
# # pprint(t.conflict,expand_all=True)
# # print("[+] File saved >> conflicts.json\n")
# t.gui()
